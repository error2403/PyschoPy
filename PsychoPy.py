### required dependencies ###
# pip install pygame
# pip install pygame_widgets
# pip install openpyxl

# python version 3.11.9

import pygame
import pygame_widgets
from pygame_widgets.slider import Slider
from pygame_widgets.textbox import TextBox
import random
import time
import os
import openpyxl
from typing import List, Dict


class Audio:
    def __init__(self):
        pygame.init()
        pygame.mixer.init()

    def play(self, file: str, loops: int = 0):
        pygame.mixer.music.load(file)
        pygame.mixer.music.play(loops)

    def stop(self):
        pygame.mixer.music.stop()

    def volchange(self, volume: float):
        pygame.mixer.music.set_volume(volume)  # The set_volume range is from 0.00 to 1.00 (every 0.01)


# game constants
IS_TWO_OPTIONS = False     # this swaps between 2 option trials and 4 option trials

    # NUM_FOLDER_X_TRIALS ex: [12, 7] --> pull 12 trials from first folder and 7 trials from second folder
if IS_TWO_OPTIONS:
    NUM_FOLDER_X_TRIALS = [2,1]
    RESULTS_FILE = 'results_2_option.xlsx'
else:
    NUM_FOLDER_X_TRIALS = [2,1,1,1]
    RESULTS_FILE = 'results_4_option.xlsx'

IMGAGE_SCALING = (5,5)
NUM_TRIALS = sum(NUM_FOLDER_X_TRIALS)
ALLOW_DUPLICATES = False

# font constants
DARK_TEAL = (21, 102, 105)
PINK = (255, 153, 153)
GRAY = (128, 128, 128)
WHITE = (255, 255, 255)
FONT_SIZE = 132

# stats
accuracy = 0
trials_completed = 0
average_reaction_time = 0

# global game variables
key_mapping: Dict[int, int] = {}
screen: pygame.Surface
screen_size: tuple = (int, int)
text_1: pygame.Surface      # the number "1" that displays under trial images
text_2: pygame.Surface      # the number "2" that displays under trial images
text_3: pygame.Surface      # the number "3" that displays under trial images
text_4: pygame.Surface      # the number "4" that displays under trial images
text_L: pygame.Surface      # the number "L" that displays under trial images
text_R: pygame.Surface      # the number "R" that displays under trial images
hands: pygame.Surface       # the image of hands
trial_directory: str
all_trials: List[Dict[str, List[str]]] = []
selected_trials: List[tuple] = []
audio_player = Audio()
patient_id = "TAG"


def initialize():
    """
    run initialization for game.
    """
    # load global variables
    global key_mapping
    global screen
    global screen_size
    global text_1, text_2, text_3, text_4, text_L, text_R
    global hands
    global trial_directory
    global all_trials
    global selected_trials
    global patient_id

    pygame.init()

    # configure keys
    if IS_TWO_OPTIONS:
        key_mapping = {pygame.K_LEFT: 1, pygame.K_RIGHT: 2}
    else:
        key_mapping = {pygame.K_1: 1, pygame.K_2: 2, pygame.K_3: 3, pygame.K_4: 4}

    # set up screen
    screen = pygame.display.set_mode((0,0), pygame.FULLSCREEN)
    pygame.display.set_caption("Audio-Visual Experiment")
    screen_size = pygame.display.get_window_size()
    screen.fill(WHITE)

    # create number fonts to go under images
    font_1 = pygame.font.SysFont('1', FONT_SIZE)
    text_1 = font_1.render('1', True, DARK_TEAL)

    font_2 = pygame.font.SysFont('2', FONT_SIZE)
    text_2 = font_2.render('2', True, DARK_TEAL)

    font_3 = pygame.font.SysFont('3', FONT_SIZE)
    text_3 = font_3.render('3', True, DARK_TEAL)

    font_4 = pygame.font.SysFont('4', FONT_SIZE)
    text_4 = font_4.render('4', True, DARK_TEAL)

    font_L = pygame.font.SysFont('L', FONT_SIZE)
    text_L = font_L.render('L', True, DARK_TEAL)

    font_R = pygame.font.SysFont('R', FONT_SIZE)
    text_R = font_R.render('R', True, DARK_TEAL)

    # load hand image
    hands = pygame.image.load("hands.jpg")
    hands = pygame.transform.scale(hands, (screen_size[0]/IMGAGE_SCALING[0], screen_size[1]/IMGAGE_SCALING[1]))
    hands.convert()

    # get directory
    game_directory = os.path.dirname(os.path.realpath(__file__))

    # load trials
    if IS_TWO_OPTIONS:
        trial_directory = game_directory + "\\2_option_trials"
    else:
        trial_directory = game_directory + "\\4_option_trials"

    trial_data = os.listdir(trial_directory)

    for item in trial_data:
        if os.path.isdir(trial_directory + f"\\{item}"):
            print("is dir: " + item)
            trial_dict: Dict[str, List[str]] = {}
            trial_dict[item] = os.listdir(trial_directory + f"\\{item}")
            all_trials.append(trial_dict)
        else:
            print("is not dir: " + item)

    # generate list of trials to run
    for folder_num, num_trials in enumerate(NUM_FOLDER_X_TRIALS):
        folder_dict = all_trials[folder_num]
        break_dict = list(folder_dict.items())[0]
        folder_name = break_dict[0]
        folder_trials = break_dict[1]

        # check if program is asking for more trials than are available
        if not ALLOW_DUPLICATES:
            if num_trials > len(folder_trials):
                print(f"Program does not allow duplicates. Asking for {num_trials} trials when there are {len(folder_trials)} available in {folder_name}")
                exit(10)
        
        # pull random trials from folder
        for _ in range(num_trials):
            temp_trial = random.choice(folder_trials)

            if not ALLOW_DUPLICATES:
                folder_trials.remove(temp_trial)
            
            selected_trials.append((folder_name, temp_trial))

    # randomize list
    random.shuffle(selected_trials)

    # create results file if not already made
    dir_contents = os.listdir()
    if dir_contents.count(RESULTS_FILE) == 0:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "summary"
        headers = ["Participant ID", "Avg Reaction Experimental", "Accuracy Experimental", "Avg Reaction Control", "Accuracy Control"]
        worksheet.append(headers)
        workbook.save(RESULTS_FILE)


def thank_you():
    """
    display the thank you screen and wait for user to press space.
    """
    # load global variables
    global screen
    global trial_directory

    # display thank you screen
    img = pygame.image.load(trial_directory + "\\thank_you.png")
    img = pygame.transform.scale(img, (screen_size[0], screen_size[1]))
    img.convert()
    
    screen.fill(WHITE)
    screen.blit(img, (0,0))
    pygame.display.update()

    # wait for user to press space
    response = None
    while response != pygame.K_SPACE:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                exit()
            elif event.type == pygame.KEYDOWN:
                response = event.key

    # clear screen
    screen.fill(WHITE)


def audio_tuning():
    """
    screen where users can adjust audio levels before starting.
    press space to continue to next screen.
    """
    # load global variables
    global screen
    global audio_player
    global trial_directory

    # configure audio tuning screen
    font = pygame.font.SysFont(None, FONT_SIZE)
    text = font.render('Audio Tuning', True, PINK)
    text_rect = text.get_rect(center=(screen_size[0]/2, screen_size[1]/3))

    font2 = pygame.font.SysFont(None, int(FONT_SIZE/2))
    text2 = font2.render('Press Space to continue', True, GRAY)
    text2_rect = text2.get_rect(center=(screen_size[0]/2, 5*screen_size[1]/6))

    slider = Slider(screen,
                    screen_size[0]/2 - screen_size[0]/6,
                    int(4*screen_size[1]/7),
                    screen_size[0]/3,
                    40,
                    min=0,
                    max=100,
                    step=1)
    output = TextBox(screen,
                     int(screen_size[0]/2) - 25,
                     int(4.5*screen_size[1]/7),
                     50,
                     50,
                     fontSize=30)
    output.disable()  # Act as label instead of textbox

    # loop audio sounds
    audio_player.play(trial_directory + '\\audio_tuning.wav', -1)

    # wait for user to press space
    response = None
    while response != pygame.K_SPACE:
        events = pygame.event.get()
        for event in events:
            if event.type == pygame.QUIT:
                pygame.quit()
                exit()
            elif event.type == pygame.KEYDOWN:
                response = event.key

        screen.fill(WHITE)
        screen.blit(text, text_rect)
        screen.blit(text2, text2_rect)

        output.setText(slider.getValue())
        audio_player.volchange(slider.getValue()/100)
        pygame_widgets.update(events)
        pygame.display.update()

    # clear screen
    screen.fill(WHITE)
    audio_player.stop()


def instructions():
    """
    display instructions for how to play game.
    press space to continue to next screen.
    """
    # load global variables
    global screen
    global trial_directory

    # display instructions screen
    img = pygame.image.load(trial_directory + "\\instructions.png")
    img = pygame.transform.scale(img, (screen_size[0], screen_size[1]))
    img.convert()
    
    screen.fill(WHITE)
    screen.blit(img, (0,0))
    pygame.display.update()

    # wait for user to press space
    response = None
    while response != pygame.K_SPACE:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                exit()
            elif event.type == pygame.KEYDOWN:
                response = event.key

    # clear screen
    screen.fill(WHITE)


def run_trials():
    if IS_TWO_OPTIONS:
        run_2_option_trials()
    else:
        run_4_option_trials()


def run_2_option_trials():
    """
    runs the PsychoPy 2 option test trials.
    """
    # load global variables
    global key_mapping
    global screen
    global screen_size
    global text_L, text_R
    global hands
    global trial_directory
    global selected_trials
    global accuracy
    global trials_completed
    global average_reaction_time
    global audio_player
    global patient_id

    # set up patient result saving location
    workbook = openpyxl.load_workbook(RESULTS_FILE)
    worksheet = workbook['summary']

    patient_id_num = 0
    for row in worksheet.values:
        patient_id_num += 1
    patient_id += f"{patient_id_num}"
    workbook.create_sheet(f"{patient_id}")

    worksheet = workbook[f"{patient_id}"]
    headers = ["Trial", "Reaction Time", "Accuracy"]
    worksheet.append(headers)             
        
    ## loop for NUM_TRIALS
    for trial_num in range(NUM_TRIALS):
        # get next trial in list
        trial = selected_trials[trial_num]

        # load trial data
        trial_data = os.listdir(trial_directory + f"\\{trial[0]}\\{trial[1]}")
        images: List[str] = []
        audio: List[str] = []
        other: List[str] = []
        for file in trial_data:
            if file.endswith('.jpg') or file.endswith('.png'):
                images.append(file)
            elif file.endswith('.wav'):
                audio.append(file)
            else:
                other.append(file)

        # randomize image array
        random.shuffle(images)

        # determine correct image
        correct_key = -1
        for image in images:
            if image.__contains__("1"):
                correct_key = images.index(image) + 1

        # place images on screen
        converted_images = []
        for image in images:
            img = pygame.image.load(trial_directory + f"\\{trial[0]}\\{trial[1]}\\{image}")
            img = pygame.transform.scale(img, (screen_size[0]/IMGAGE_SCALING[0], screen_size[1]/IMGAGE_SCALING[1]))
            img.convert()
            converted_images.append(img)

        screen.fill(WHITE)
        screen.blit(converted_images[0], (7*screen_size[0]/25, screen_size[1]/3))
        screen.blit(converted_images[1], (13*screen_size[0]/25, screen_size[1]/3))

        screen.blit(text_L, (9.3*screen_size[0]/25, 1.75*screen_size[1]/3))
        screen.blit(text_R, (15.3*screen_size[0]/25, 1.75*screen_size[1]/3))

        # display screen
        pygame.display.update()

        # play audio file
        audio_player.play(trial_directory + f"\\{trial[0]}\\{trial[1]}\\{audio[0]}")

        # Record start time
        start_time = time.time()

        # wait for user input
        response = None
        while response not in key_mapping:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()
                elif event.type == pygame.KEYDOWN:
                    response = event.key

        # record reaction time
        reaction_time = time.time() - start_time
        average_reaction_time += reaction_time

        # check for correct selection
        selected_image = key_mapping.get(response)
        if selected_image == correct_key:
            answer = "correct"
            accuracy += 1
        else:
            answer = "incorrect"
            accuracy += 0
        trials_completed += 1

        # display feedback
        print(f"Trial {trial} - Reaction Time: {reaction_time:.2f}s, {answer}")
        data = [f"{trial}", f"{reaction_time:.2f}s", f"{answer}"]
        worksheet.append(data)

        # stop audio if still playing
        audio_player.stop()

        # 2 second hand delay thing
        screen.fill(WHITE)
        hands_rect = hands.get_rect(center=(screen_size[0]/2, screen_size[1]/2))
        screen.blit(hands, hands_rect)
        pygame.display.update()
        time.sleep(2)

    ## end loop for NUM_TRIALS
    workbook.save(RESULTS_FILE)


def run_4_option_trials():
    """
    runs the PsychoPy 4 option test trials.
    """
    # load global variables
    global key_mapping
    global screen
    global screen_size
    global text_1, text_2, text_3, text_4
    global hands
    global trial_directory
    global selected_trials
    global accuracy
    global trials_completed
    global average_reaction_time
    global audio_player
    global patient_id

    # set up patient result saving location
    workbook = openpyxl.load_workbook(RESULTS_FILE)
    worksheet = workbook['summary']

    patient_id_num = 0
    for row in worksheet.values:
        patient_id_num += 1
    patient_id += f"{patient_id_num}"
    workbook.create_sheet(f"{patient_id}")

    worksheet = workbook[f"{patient_id}"]
    headers = ["Trial", "Reaction Time", "Accuracy"]
    worksheet.append(headers)             
        
    ## loop for NUM_TRIALS
    for trial_num in range(NUM_TRIALS):
        # get next trial in list
        trial = selected_trials[trial_num]

        # load trial data
        trial_data = os.listdir(trial_directory + f"\\{trial[0]}\\{trial[1]}")
        images: List[str] = []
        audio: List[str] = []
        other: List[str] = []
        for file in trial_data:
            if file.endswith('.jpg') or file.endswith('.png'):
                images.append(file)
            elif file.endswith('.wav'):
                audio.append(file)
            else:
                other.append(file)

        # randomize image array
        random.shuffle(images)

        # determine correct image
        correct_key = -1
        for image in images:
            if image.__contains__("1"):
                correct_key = images.index(image) + 1

        # place images on screen
        converted_images = []
        for image in images:
            img = pygame.image.load(trial_directory + f"\\{trial[0]}\\{trial[1]}\\{image}")
            img = pygame.transform.scale(img, (screen_size[0]/IMGAGE_SCALING[0], screen_size[1]/IMGAGE_SCALING[1]))
            img.convert()
            converted_images.append(img)

        screen.fill(WHITE)
        screen.blit(converted_images[0], (screen_size[0]/25, screen_size[1]/3))
        screen.blit(converted_images[1], (7*screen_size[0]/25, screen_size[1]/3))
        screen.blit(converted_images[2], (13*screen_size[0]/25, screen_size[1]/3))
        screen.blit(converted_images[3], (19*screen_size[0]/25, screen_size[1]/3))

        screen.blit(text_1, (3.3*screen_size[0]/25, 1.75*screen_size[1]/3))
        screen.blit(text_2, (9.3*screen_size[0]/25, 1.75*screen_size[1]/3))
        screen.blit(text_3, (15.3*screen_size[0]/25, 1.75*screen_size[1]/3))
        screen.blit(text_4, (21.3*screen_size[0]/25, 1.75*screen_size[1]/3))

        # display screen
        pygame.display.update()

        # play audio file
        audio_player.play(trial_directory + f"\\{trial[0]}\\{trial[1]}\\{audio[0]}")

        # Record start time
        start_time = time.time()

        # wait for user input
        response = None
        while response not in key_mapping:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()
                elif event.type == pygame.KEYDOWN:
                    response = event.key

        # record reaction time
        reaction_time = time.time() - start_time
        average_reaction_time += reaction_time

        # check for correct selection
        selected_image = key_mapping.get(response)
        if selected_image == correct_key:
            answer = "correct"
            accuracy += 1
        else:
            answer = "incorrect"
            accuracy += 0
        trials_completed += 1

        # display feedback
        print(f"Trial {trial} - Reaction Time: {reaction_time:.2f}s, {answer}")
        data = [f"{trial}", f"{reaction_time:.2f}s", f"{answer}"]
        worksheet.append(data)

        # stop audio if still playing
        audio_player.stop()

        # 2 second hand delay thing
        screen.fill(WHITE)
        hands_rect = hands.get_rect(center=(screen_size[0]/2, screen_size[1]/2))
        screen.blit(hands, hands_rect)
        pygame.display.update()
        time.sleep(2)

    ## end loop for NUM_TRIALS
    workbook.save(RESULTS_FILE) 

def clean_up():
    """
    
    """
    # load global constants
    global average_reaction_time
    global accuracy
    global trials_completed
    global patient_id

    # display game stats
    average_reaction_time /= NUM_TRIALS
    print(f"Stats - Average Reaction Time: {average_reaction_time:.2f}s, Accuracy: {accuracy}/{trials_completed}")

    # write data to xlsx file
    workbook = openpyxl.load_workbook(RESULTS_FILE)
    worksheet = workbook['summary']
    data = [f"{patient_id}", f"{average_reaction_time:.2f}s", f"{accuracy}/{trials_completed}", f"{average_reaction_time:.2f}s", f"{accuracy}/{trials_completed}"]
    worksheet.append(data)
    workbook.save(RESULTS_FILE)
    
    # kill pygame
    pygame.quit()


def main():
    """
    overall structure for game.
    """
    initialize()
    thank_you()
    audio_tuning()
    instructions()
    run_trials()
    clean_up()


if __name__ == "__main__":
    main()